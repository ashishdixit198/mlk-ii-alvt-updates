# -*- coding: utf-8 -*-
import sys
import re
import io
import math
import json
import csv
import os
import hashlib
import zipfile
import tempfile
import shutil
from dataclasses import dataclass
from typing import List, Tuple, Dict, Any, Optional, Iterable, Set, Callable
from collections import Counter, defaultdict
from datetime import datetime, timedelta, time as dtime
from pathlib import Path

# Toolhub Imports
try:
    from core.communication_link_analysis import CommunicationLinkAnalyzer
except ImportError:
    # Fallback for relative imports if run as a script in the core directory
    try:
        from communication_link_analysis import CommunicationLinkAnalyzer
    except ImportError:
        CommunicationLinkAnalyzer = None

# Excel / Workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import FormulaRule

# Qt
from PySide6.QtCore import QThread, Signal

# Matplotlib
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

# ---- Constants ----
APP_ID = "LOGTK"

# ---- EVENT/USER Log Regexes ----
ENTRY_RE = re.compile(r'^\[(event|warning|error)\]\t([\d./ -]+[0-9:. ]+[AP]M)\t(.*?\t\t)', re.MULTILINE | re.IGNORECASE)
HDR_PROGRAM_RE = re.compile(r'Program\s+"([^"]+)"\s+at address\s+(\d+)\s+on\s+(\S+)', re.IGNORECASE)
HDR_LINE_KEYS = ["Executive Version", "Executive CRC", "Compiler Version", "Application Version", "Application CRC", "Maintenance Tool Version", "PC Date"]

# Common User/Bit Timestamps (Flexible)
TS_RE_DEFAULT = re.compile(r'^\(?(?P<date>\d{1,2}[./-]\d{1,2}[./-]\d{2,4})\)?\s+\(?(?P<time>\d{1,2}:\d{2}:\d{2}(?:\.\d+)?)\)?(?:\s+\(?(?P<ampm>[AP]M)?\)?)?$', re.IGNORECASE)
BIT_RE = re.compile(r'^(?P<name>.*?)(?:\s*\(?\s*bit(?:\s+no\.?)?\s+(?P<no>\d+)\s*\)?)\s+is\s+(?P<st>SET|CLEAR)\b.*$', re.IGNORECASE)
PROGRAM_RE = re.compile(r'^Program\s+"([^"]+)"', re.IGNORECASE)
EXEC_CRC_RE = re.compile(r'^\s*Executive\s+CRC(?:\s+is|:)?\s*([0-9A-Fa-f]+)\b', re.IGNORECASE)
APP_CRC_RE = re.compile(r'^\s*Application\s+CRC(?:\s+is|:)?\s*([0-9A-Fa-f]+)\b', re.IGNORECASE)

ERR_HDR_PROGRAM_RE = re.compile(r'Program\s+"(?P<program>[^"]+)"\s+at address\s+(?P<address>\d+)\s+on\s+(?P<com>\S+)', re.IGNORECASE)
ERR_PC_DATE_RE = re.compile(r'PC Date\s*:(?P<pcdate>.*)', re.IGNORECASE)
YEAR_CHANGE_RE = re.compile(r"Year is changing from '?(?P<from_y>\d+) to '?(?P<to_y>\d+)", re.IGNORECASE)
# Extremely robust regex: handles single/double digit hours, optional leading spaces, AM/PM, and optional codes (with or without brackets).
ERR_ENTRY_RE = re.compile(r'^\[(?P<level>ERROR|EVENT)\]\s+(?P<md>[\d./-]+)\s+(?P<time>\s*\d{1,2}:\d{2}:\d{2}(?:\.\d+)?)\s+(?P<ampm>AM|PM)\s+(?P<message>.*?)(?:\s+\[?(?P<codes>(?:[0-9A-Fa-f]{4}\s*)+)\]?)?\s*$', re.IGNORECASE)

# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------

def _safe_temp_root() -> Path:
    return Path(tempfile.mkdtemp(prefix=f"{APP_ID}_"))

def _extract_logs_from_zip(zip_path: Path, out_dir: Path, name_filter_substring: Optional[str]=None) -> list[Path]:
    extracted: list[Path] = []
    try:
        with zipfile.ZipFile(zip_path, 'r') as zf:
            for info in zf.infolist():
                if info.is_dir(): continue
                name = Path(info.filename).name
                if not name.lower().endswith('.log'): continue
                if name_filter_substring and name_filter_substring not in name.lower(): continue
                dest = out_dir / name
                with zf.open(info, 'r') as src, open(dest, 'wb') as dst:
                    shutil.copyfileobj(src, dst)
                extracted.append(dest)
    except Exception as ex:
        # Failed to extract from ZIP
        pass
    return extracted

def autoformat(ws: Worksheet, freeze_header: bool = True, sample_rows: int = 80):
    ws.auto_filter.ref = ws.dimensions
    if freeze_header: ws.freeze_panes = "A2"
    max_col = ws.max_column
    max_row = min(ws.max_row, sample_rows)
    for c in range(1, max_col + 1):
        max_len = 10
        for r in range(1, max_row + 1):
            v = ws.cell(r, c).value
            if v is None: continue
            l = len(str(v))
            if l > max_len: max_len = l
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 60)

# -----------------------------------------------------------------------------
# EVENT Logic
# -----------------------------------------------------------------------------


HDR_LINE_KEYS = ["Executive Version", "Executive CRC", "Compiler Version", "Application Version", "Application CRC", "Maintenance Tool Version", "PC Date"]

@dataclass
class Rule:
    name: str; label: str; regex: re.Pattern; fields: List[str]; timeline: bool
    @staticmethod
    def from_dict(d: Dict[str, Any]) -> 'Rule':
        return Rule(name=d['name'], label=d.get('label', d['name']), regex=re.compile(d['regex'], re.IGNORECASE), fields=list(d.get('fields', [])), timeline=bool(d.get('timeline', True)))

def default_rules() -> List[Rule]:
    raw = [
        {"name": "ack_timeouts", "label": "ACK timeouts (all)", "regex": r"MII Peer - Acknowledge Timeout Serial link (?P<Link>\S+), Address (?P<Address>\d+)", "fields": ["Link", "Address"], "timeline": True},
        {"name": "station_down", "label": "Station Down events", "regex": r"MII Peer - Station Down - Stale Data Timeout Serial link (?P<Link>\S+), Address (?P<Address>\d+)", "fields": ["Link", "Address"], "timeline": True},
        {"name": "station_up", "label": "Station Up events", "regex": r"MII Peer - Station Up Serial link (?P<Link>\S+), Address (?P<Address>\d+)", "fields": ["Link", "Address"], "timeline": True},
        {"name": "genisys_fail", "label": "GENISYS failures", "regex": r"GENISYS slave: slave communication failed - Serial link (?P<SerialLink>\S+), Address (?P<Address>\d+)", "fields": ["SerialLink", "Address"], "timeline": True},
        {"name": "free_pool_low", "label": "Free pool low events", "regex": r"New low count for free pool buffer list - Link (?P<Link>\S+), Count = (?P<Count>\d+)", "fields": ["Link", "Count"], "timeline": True},
    ]
    return [Rule.from_dict(x) for x in raw]

def load_rules(rules_path: Optional[str]) -> List[Rule]:
    if not rules_path: return default_rules()
    try:
        p = Path(rules_path).expanduser()
        with p.open('r', encoding='utf-8') as f: return [Rule.from_dict(x) for x in json.load(f)]
    except Exception: return default_rules()

def message_signature(msg: str) -> str:
    sig = re.sub(r"\d+", "<#>", msg)
    sig = re.sub(r"\s+", " ", sig).strip()
    return sig[:200]




def process_event_files(files: List[Path], out_dir: Path, rules: List[Rule], progress_cb: Optional[Callable[[int, int], None]] = None, cancel_check: Optional[Callable[[], bool]] = None, sort_chronologically: bool = True, log_callback: Optional[Callable[[str], None]] = None, filter_types: Optional[Set[str]] = None, logic_file_path: Optional[str] = None) -> Tuple[Optional[Path], int, Dict[str, Any]]:
    if filter_types is None:
        filter_types = {'event', 'warning', 'error'}  # Default to all types
    
    # Logic Cache: {program_name: [station_names]}
    logic_cache: Dict[str, List[str]] = {}
    # File mapping: {log_name: {"program": prog, "logic_file": logic_path}}
    file_mapping: Dict[str, Dict[str, str]] = {}
    logic_source = Path(logic_file_path) if logic_file_path else None
    
    def get_station_names(prog: str, log_name: str) -> List[str]:
        if not logic_source or not CommunicationLinkAnalyzer: return []
        if prog in logic_cache: 
            # Record mapping from existing cache entry
            # Finding which logic file was previously assigned to this program
            for l_name, m_info in file_mapping.items():
                if m_info["program"] == prog:
                    file_mapping[log_name] = {"program": prog, "logic_file": m_info["logic_file"]}
                    break
            return logic_cache[prog]
        
        target_file = None
        if logic_source.is_file():
            target_file = logic_source
        elif logic_source.is_dir():
            # Search for matching logic file
            for ext in (".txt", ".mll", ".ML2", ".txt.log"):
                potential = logic_source / f"{prog}{ext}"
                if potential.exists():
                    target_file = potential
                    break
            if not target_file:
                # Fuzzy search (RECURSIVE) - Supports .txt or .ml2
                for f in logic_source.rglob("*"):
                    if f.is_file() and prog.upper() in f.name.upper() and f.suffix.lower() in (".txt", ".ml2"):
                        target_file = f
                        break
        
        if target_file:
            try:
                if log_callback: log_callback(f"Loading logic for {prog}: {target_file.name}")
                analyzer = CommunicationLinkAnalyzer(None)
                links = analyzer.extract_comm_links(str(target_file))
                names = [link.get('Station Name', 'UNKNOWN') for link in links]
                logic_cache[prog] = names
                file_mapping[log_name] = {"program": prog, "logic_file": target_file.name}
                return names
            except Exception as e:
                if log_callback: log_callback(f"Error loading {target_file.name}: {str(e)}")
        
        file_mapping[log_name] = {"program": prog, "logic_file": "NOT FOUND"}
        return []

    combined_entries: List[Dict[str, Any]] = []
    headers_rows: List[List[Any]] = []
    
    for i, f in enumerate(files, start=1):
        if cancel_check and cancel_check(): break
        if progress_cb: progress_cb(i, len(files))
        if log_callback: log_callback(f"Parsing: {f.name}")
        
        text = Path(f).read_text(encoding='utf-8', errors='ignore')
        header = EventWorker._parse_header(text)
        program = header.get("Program", "Unknown")
        year = EventWorker._infer_year(header)
        
        # Load Station Map for this specific program
        current_stations = get_station_names(str(program), f.name)
        
        # Detect System (A/B)
        # Rule 1: First character of filename
        f_name_upper = f.name.upper()
        if f_name_upper.startswith('A'): system = "A"
        elif f_name_upper.startswith('B'): system = "B"
        else: system = "Unknown"
        
        # Rule 2: Folder or Content hints if still Unknown
        if system == "Unknown":
            txt_upper = text[:3000].upper()
            p_name = f.parent.name.upper()
            if "_A" in f_name_upper or "SYSTEM A" in txt_upper or "_A" in p_name or "SYSTEM A" in p_name: system = "A"
            elif "_B" in f_name_upper or "SYSTEM B" in txt_upper or "_B" in p_name or "SYSTEM B" in p_name: system = "B"
        
        if system == "Unknown":
            # Rule 3: Check Site ID or other header hints
            site_id = header.get("Site ID", "").upper()
            if "A" in site_id: system = "A"
            elif "B" in site_id: system = "B"
            
        if system == "Unknown" and "address" in str(header.get("Program", "")).lower():
            # Rule 4: Program string hints
            prog_str = str(header.get("Program", "")).upper()
            if "SYSTEM A" in prog_str: system = "A"
            elif "SYSTEM B" in prog_str: system = "B"

        # Combine for full system name
        system_full = f"{program}_{system}"
        
        headers_rows.append([f.parent.name, f.name, header.get("Program", ""), header.get("Executive Version", ""), header.get("Executive CRC", ""), header.get("Application CRC", ""), header.get("PC Date", ""), system_full])
        
        for line in text.splitlines():
            myc = YEAR_CHANGE_RE.search(line)
            if myc:
                new_y = int(myc.group('from_y'))
                year = new_y + 2000 if new_y < 100 else new_y

            m_entry = ENTRY_RE.match(line)
            if not m_entry:
                continue
            
            level, ts, msg_part = m_entry.group(1), m_entry.group(2), m_entry.group(3)
            # Entry RE usually captures up to double tab. Msg might have trailing tabs.
            msg = msg_part.strip()
            
            # Extract Hex codes (everything after double tab in the original line)
            hex_match = re.search(r'\t\t([0-9A-Fa-f\s]{16,24})$', line)
            hex_codes = hex_match.group(1).strip() if hex_match else ""

            if level.lower() not in filter_types:
                continue
            
            dt = None
            for fmt in ("%m/%d %I:%M:%S.%f %p", "%m/%d %I:%M:%S %p"):
                try: dt = datetime.strptime(ts, fmt).replace(year=year); break
                except Exception: pass
            
            entry = {
                "level": level,
                "dt": dt,
                "dt_display": dt.strftime("%Y-%m-%d %H:%M:%S.%f")[:-5] if dt else "",
                "msg": msg,
                "hex": hex_codes,
                "file_name": f.name,
                "parent_name": f.parent.name,
                "system": system_full,
                "station_name": ""
            }

            # Try to enrich Station Name if Address is in message
            addr_match = re.search(r'Address\s+(\d+)\b', msg, re.IGNORECASE)
            if addr_match:
                addr_num = int(addr_match.group(1))
                if 0 <= addr_num < len(current_stations):
                    entry["station_name"] = current_stations[addr_num]
            
            combined_entries.append(entry)

    if not combined_entries: return None, 0, {}
    
    if sort_chronologically:
        combined_entries.sort(key=lambda x: x["dt"] or datetime.min)
    
    rule_counts: Dict[str, Counter] = defaultdict(Counter)
    simple_totals = Counter()
    tl = defaultdict(lambda: defaultdict(int))
    unclassified = Counter()
    unclassified_examples = {}
    comm_events = []

    for entry in combined_entries:
        msg = entry["msg"]
        dt = entry["dt"]
        matched = False
        for r in rules:
            mm = r.regex.search(msg)
            if mm:
                matched = True
                if r.fields: 
                    fields_tuple = tuple(mm.group(f) for f in r.fields)
                    rule_counts[r.name][fields_tuple] += 1
                else: 
                    simple_totals[r.name] += 1
                
                if dt and r.timeline: 
                    tl[dt.replace(minute=0, second=0, microsecond=0)][r.name] += 1
                break
        
        if not matched:
            sig = message_signature(msg)
            unclassified[sig] += 1
            if sig not in unclassified_examples: unclassified_examples[sig] = msg
        
        if dt: 
            tl[dt.replace(minute=0, second=0, microsecond=0)]["total_entries"] += 1
        
        if entry["station_name"] or "Address" in msg:
            comm_events.append(entry)

    overview_rows = [("Files processed", len(files)), ("Total entries", len(combined_entries)), ("Comm Link Events", len(comm_events))]
    label_map = {r.name: r.label for r in rules}
    for name, ctr in sorted(rule_counts.items()): overview_rows.append((label_map.get(name, name), int(sum(ctr.values()))))
    for name, cnt in sorted(simple_totals.items()): overview_rows.append((label_map.get(name, name), int(cnt)))
    
    out_path = out_dir / f"Event_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # USE THE COMPREHENSIVE WORKBOOK WRITING
    EventWorker._write_static_workbook_enhanced(
        out_path, 
        overview_rows, 
        headers_rows, 
        tl, 
        rules, 
        combined_entries, 
        unclassified, 
        unclassified_examples,
        comm_events
    )
    
    return out_path, len(combined_entries), {"comm_events": comm_events, "mapping": file_mapping}

class EventWorker(QThread):
    log_msg = Signal(str); progress = Signal(int, int); result_ready = Signal(dict); failed = Signal(str)
    def __init__(self, mode: str, selection: List[str], out_dir: str, rules_path: Optional[str], export_proposed: bool, sort_chronologically: bool = True, filter_types: Tuple[str, ...] = ('event', 'warning', 'error'), logic_file: Optional[str] = None):
        super().__init__()
        self.mode, self.selection, self.out_dir = mode, selection, Path(out_dir)
        self.rules_path, self.export_proposed = rules_path, export_proposed
        self.sort_chronologically = sort_chronologically
        self.filter_types = {t.lower() for t in filter_types}  # Normalize to lowercase set
        self.logic_file = logic_file
        self._cancel = False
    def request_cancel(self): self._cancel = True
    def _collect_files(self) -> List[Path]:
        _td = _safe_temp_root()
        _z = lambda p: _extract_logs_from_zip(p, _td, 'event') if p.suffix.lower() == '.zip' else []
        if self.mode == 'folder':
            f = Path(self.selection[0]); l = []
            for p in f.rglob('*'):
                if p.is_file() and 'event' in p.name.lower() and p.suffix.lower() == '.log': l.append(p)
                elif p.is_file() and p.suffix.lower() == '.zip': l.extend(_z(p))
            return sorted(l)
        if self.mode == 'files':
            l = []
            for s in self.selection:
                p = Path(s)
                if p.is_file() and p.suffix.lower() == '.zip': l.extend(_z(p))
                elif p.is_file() and 'event' in p.name.lower() and p.suffix.lower() == '.log': l.append(p)
            return sorted(l)
        if self.mode == 'single':
            p = Path(self.selection[0])
            if p.suffix.lower() == '.zip': return _extract_logs_from_zip(p, _td, 'event')
            return [p] if p.is_file() else []
        return []
    @staticmethod
    def _parse_header(text: str) -> Dict[str, Any]:
        h: Dict[str, Any] = {}
        for line in text.splitlines()[:50]:
            if line.startswith('['): break
            if 'Program' in line:
                m = HDR_PROGRAM_RE.search(line)
                if m: h["Program"], h["Address"], h["Port"] = m.group(1), int(m.group(2)), m.group(3)
            else:
                for k in HDR_LINE_KEYS:
                    if line.startswith(k): h[k] = line.split(":", 1)[1].strip(); break
        return h
    @staticmethod
    def _infer_year(header: Dict[str, Any]) -> int:
        if "PC Date" in header:
            m = re.search(r"\b(20\d{2})\b", header["PC Date"])
            if m: return int(m.group(1))
        return datetime.now().year
    @staticmethod
    def _write_static_workbook_enhanced(out_path: Path, overview_rows: List[Tuple[str, Any]], headers_rows: List[List[Any]], tl: Dict[datetime, Dict[str, int]], rules: List[Rule], entries: List[Dict[str, Any]], unclassified: Counter, unclassified_examples: Dict[str, str], comm_events: List[Dict[str, Any]]):
        wb = Workbook()
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(bold=True)

        def style_header(ws):
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

        # 1. Overview
        ws_over = wb.active; ws_over.title = "Overview"
        ws_over.append(["Field", "Value"])
        for k, v in overview_rows: ws_over.append([k, v])
        style_header(ws_over); autoformat(ws_over)

        # 2. Headers
        ws_hdr = wb.create_sheet("Headers")
        ws_hdr.append(["File Path", "File Name", "Program", "Executive Version", "Executive CRC", "Application CRC", "PC Date", "System"])
        for row in headers_rows: ws_hdr.append(row)
        style_header(ws_hdr); autoformat(ws_hdr)

        # 3. Timeline
        ws_tl = wb.create_sheet("Timeline_Hourly")
        timeline_rule_names = {r.name for r in rules if r.timeline}
        observed = set()
        for d in tl.values(): observed.update(d.keys())
        timeline_cols = sorted((timeline_rule_names & observed) - {"total_entries"})
        ws_tl.append(["Hour", "total_entries"] + timeline_cols)
        for h in sorted(tl.keys()):
            row = [h, int(tl[h].get("total_entries", 0))]
            for c in timeline_cols: row.append(int(tl[h].get(c, 0)))
            ws_tl.append(row)
        style_header(ws_tl); autoformat(ws_tl)

        # 4. Comm Link Analysis (NEW)
        ws_comm = wb.create_sheet("Comm_Link_Analysis")
        ws_comm.append(["Source File", "System", "DateTime", "Address", "Station Name", "Level", "Message", "Hex Codes"])
        for e in comm_events:
            addr_match = re.search(r'Address\s+(\d+)\b', e["msg"], re.IGNORECASE)
            addr = addr_match.group(1) if addr_match else ""
            # Use dt_display if available to avoid trailing zeros
            dt_val = e.get("dt_display", e["dt"])
            ws_comm.append([e["file_name"], e["system"], dt_val, addr, e["station_name"], e["level"], e["msg"], e["hex"]])
        style_header(ws_comm); autoformat(ws_comm)

        # 5. Raw Events
        ws_raw = wb.create_sheet("Raw_Events")
        ws_raw.append(["File Name", "System", "Level", "DateTime", "Station Name", "Message", "Hex Codes"])
        for e in entries:
            dt_val = e.get("dt_display", e["dt"])
            ws_raw.append([e["file_name"], e["system"], e["level"], dt_val, e["station_name"], e["msg"], e["hex"]])
        style_header(ws_raw); autoformat(ws_raw)

        # 6. Unclassified
        ws_unc = wb.create_sheet("Unclassified_Signatures")
        ws_unc.append(["Signature", "Count", "Example"])
        for sig, cnt in unclassified.most_common(): ws_unc.append([sig, int(cnt), unclassified_examples.get(sig, "")])
        style_header(ws_unc); autoformat(ws_unc)

        wb.save(out_path)

    def run(self):
        try:
            files = self._collect_files()
            if not files: self.failed.emit("No files to process."); return
            rules = load_rules(self.rules_path)
            self.out_dir.mkdir(parents=True, exist_ok=True)
            out_path, cnt, meta = process_event_files(files, self.out_dir, rules, self.progress.emit, lambda: self._cancel, sort_chronologically=self.sort_chronologically, log_callback=self.log_msg.emit, filter_types=self.filter_types, logic_file_path=self.logic_file)
            if out_path: 
                res = {'report_path': str(out_path), 'count': cnt}
                res.update(meta)
                self.result_ready.emit(res)
            else: self.failed.emit("No event entries found.")
        except Exception as ex: 
            import traceback
            self.failed.emit(f"{str(ex)}\n{traceback.format_exc()}")

# -----------------------------------------------------------------------------
# USER (Bit Status) Logic
# -----------------------------------------------------------------------------

TS_RE = re.compile(r'^\(?(?P<date>\d{1,2}[./-]\d{1,2}[./-]\d{2,4})\)?\s+\(?(?P<time>\d{1,2}:\d{2}:\d{2}(?:\.\d+)?)\)?(?:\s+\(?(?P<ampm>[AP]M)?\)?)?$', re.IGNORECASE)
BIT_RE = re.compile(r'^(?P<name>.*?)\s*(?:\(?\s*bit(?:\s+no\.?)?\s+(?P<no>\d+)\s*\)?\s+)?is\s+(?P<st>SET|CLEAR)\b.*$', re.IGNORECASE)
PROGRAM_RE = re.compile(r'^Program\s+"([^"]+)"', re.IGNORECASE)
EXEC_CRC_RE = re.compile(r'^\s*Executive\s+CRC(?:\s+is|:)?\s*([0-9A-Fa-f]+)\b', re.IGNORECASE)
APP_CRC_RE = re.compile(r'^\s*Application\s+CRC(?:\s+is|:)?\s*([0-9A-Fa-f]+)\b', re.IGNORECASE)
HDR_ROW = ["File Path", "File Name", "Bit Name", "Date", "Time", "Bit Status", "Bit ID", "Timestamp"]
SUMMARY_HDR = ["File Path", "File Name", "Sheet", "Program", "Executive CRC", "Application CRC", "First TS", "Last TS", "Processed At"]

def parse_bit_ts(dt_str: str, tm_str: str) -> Optional[datetime]:
    if not dt_str or not tm_str: return None
    fmts = [
        "%m/%d/%y %I:%M:%S.%f %p", "%m/%d/%y %I:%M:%S %p",
        "%d/%m/%y %I:%M:%S.%f %p", "%d/%m/%y %I:%M:%S %p",
        "%m.%d.%y %I:%M:%S.%f %p", "%m.%d.%y %I:%M:%S %p",
        "%d.%m.%y %I:%M:%S.%f %p", "%d.%m.%y %I:%M:%S %p",
        "%Y-%m-%d %H:%M:%S", # Fallback for ISO-like
    ]
    ts_str = f"{dt_str} {tm_str}"
    for fmt in fmts:
        try: return datetime.strptime(ts_str, fmt)
        except: continue
    return None

def sanitize_sheet_name(name: str, existing: Set[str]) -> str:
    s = ''.join((' ' if ch in {'/', '\\', '?', '*', '[', ']', ':'} else ch) for ch in name).strip()[:31]
    if not s: s = "Sheet"
    base, i = s, 1
    while s in existing:
        suffix = f" ({i})"
        s = (base[:31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
        i += 1
    return s

def set_sheet_meta_header(ws: Worksheet, program: str, exec_crc: str, app_crc: str):
    ws.merge_cells("A1:G1") # Merged A to G
    ws["A1"] = f"Program: {program or '-'}\nExecutive CRC: {exec_crc or '-'}\nApplication CRC: {app_crc or '-'}"
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws["A1"].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    # Increase row height for wrap text
    ws.row_dimensions[1].height = 45 
    ws.freeze_panes = "A3"

def apply_status_formatting(ws: Worksheet):
    """Apply Red/Green formatting to the Status column (Column E)."""
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    
    # Status is column 6 (F). We color entire row A-H.
    for row in range(3, ws.max_row + 1):
        cell_status = ws.cell(row=row, column=6)
        val = str(cell_status.value).strip().upper() if cell_status.value else ""
        fill = None; font = None
        if val == "SET":
            fill = green_fill; font = green_font
        elif val == "CLEAR":
            fill = red_fill; font = red_font
            
        if fill:
            for col in range(1, 9): # A to H
                c = ws.cell(row=row, column=col)
                c.fill = fill
                c.font = font




def process_bit_files(files: List[Path], out_dir: Path, filter_date: Optional[str] = None, filter_start: Optional[str] = None, filter_end: Optional[str] = None, consolidate: bool = False, progress_callback=None, cancel_check=None, sort_chronologically: bool = True, log_callback=None):
    if not files: return None
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook(); ws_summary = wb.active; ws_summary.title = "Summary"; ws_summary.append(SUMMARY_HDR)
    consolidated_ws = None
    if consolidate:
        consolidated_ws = wb.create_sheet(title="Consolidated_Logs")
        set_sheet_meta_header(consolidated_ws, "Consolidated", "N/A", "N/A")
        consolidated_ws.append(HDR_ROW)
    
    all_rows_for_consolidation = [] # To be sorted later

    f_date_obj = None
    if filter_date:
        try: # Try to parse filter_date which might be ISO (YYYY-MM-DD) or DD/MM/YY
            if '-' in filter_date: f_date_obj = datetime.strptime(filter_date, "%Y-%m-%d").date()
            else: f_date_obj = datetime.strptime(filter_date, "%d/%m/%y").date()
        except: pass

    f_start = dtime.fromisoformat(filter_start) if filter_start else None
    f_end = dtime.fromisoformat(filter_end) if filter_end else None
    rows_added = 0
    for i, f in enumerate(files, start=1):
        if cancel_check and cancel_check(): break
        if progress_callback: progress_callback(i, len(files))
        if log_callback: log_callback(f"Scanning: {f.name}")
        s_name = sanitize_sheet_name(Path(f).name, {ws.title for ws in wb.worksheets})
        ws = None if consolidate else wb.create_sheet(title=s_name)
        if ws:
            set_sheet_meta_header(ws, "", "", "")
            ws.append(HDR_ROW)
        meta, cur_date_obj, cur_time, first_ts, last_ts, pending = {"program": "", "exec_crc": "", "app_crc": ""}, None, "", "", "", []
        with Path(f).open('r', encoding='utf-8', errors='replace') as fh:
            for line in (l.strip() for l in fh if l.strip()):
                if cancel_check and cancel_check(): break
                changed = False
                if not meta["program"]: m = PROGRAM_RE.match(line); (meta.update({"program": m.group(1).strip()}), changed := True) if m else None
                if not meta["exec_crc"]: m = EXEC_CRC_RE.match(line); (meta.update({"exec_crc": m.group(1).upper()}), changed := True) if m else None
                if not meta["app_crc"]: m = APP_CRC_RE.match(line); (meta.update({"app_crc": m.group(1).upper()}), changed := True) if m else None
                if changed and ws: set_sheet_meta_header(ws, meta["program"], meta["exec_crc"], meta["app_crc"])
                mts = TS_RE.match(line)
                if mts:
                    l_date, l_time = mts.group('date'), (f"{mts.group('time')} {mts.group('ampm') or ''}").strip()
                    try:
                        for fmt in ("%m/%d/%y", "%d/%m/%y", "%m.%d.%y", "%d.%m.%y", "%m-%d-%y", "%d-%m-%y", "%Y-%m-%d"):
                            try:
                                cur_date_obj = datetime.strptime(l_date, fmt).date()
                                break
                            except: continue
                    except: pass
                    cur_time = l_time
                    skip = False
                    if f_date_obj and cur_date_obj and cur_date_obj != f_date_obj: skip = True
                    if not skip and (f_start or f_end):
                        try:
                            t_part = mts.group('time')
                            t_obj = datetime.strptime(l_time, "%I:%M:%S.%f %p" if '.' in t_part else "%I:%M:%S %p").time()
                            if f_start and t_obj < f_start: skip = True
                            if f_end and t_obj > f_end: skip = True
                        except: pass
                    if not skip:
                        l_date_norm = cur_date_obj.strftime("%d/%m/%y") if cur_date_obj else l_date
                        for b_n, b_no, st_t in pending:
                            try: b_val = int(b_no) if b_no else 0
                            except: b_val = 0
                            row_data = [f.parent.name, f.name, b_n, l_date_norm, l_time, 'SET' if st_t.lower() == 'set' else 'clear', b_val, ""]
                            if ws: ws.append(row_data)
                            if consolidate: all_rows_for_consolidation.append(row_data)
                            rows_added += 1
                        ts = f"{l_date} {l_time}".strip()
                        if ts: (first_ts := ts) if not first_ts else None; last_ts = ts
                    pending.clear(); continue
                mb = BIT_RE.match(line)
                if mb:
                    pending.append((mb.group('name').strip(), mb.group('no') or '', mb.group('st')))
                    continue
            
            if pending:
                l_date_str = cur_date_obj.strftime("%d/%m/%y") if cur_date_obj else (l_date if 'l_date' in locals() else "")
                l_time_str = cur_time if cur_time else ""
                for b_n, b_no, st_t in pending:
                    try: b_val = int(b_no) if b_no else 0
                    except: b_val = 0
                    row_data = [f.parent.name, f.name, b_n, l_date_str, l_time_str, 'SET' if st_t.lower() == 'set' else 'clear', b_val, ""]
                    if ws: ws.append(row_data)
                    if consolidate: all_rows_for_consolidation.append(row_data)
                    rows_added += 1
        if ws:
            apply_status_formatting(ws)
            autoformat(ws)
        ws_summary.append([f.parent.name, f.name, s_name if not consolidate else "Consolidated", meta["program"], meta["exec_crc"], meta["app_crc"], first_ts, last_ts, datetime.now().isoformat()])
    
    # Final Sorting and Writing to Consolidated Sheet
    if consolidate and consolidated_ws:
        if sort_chronologically:
            # Sort by Date/Time columns (indices 3 and 4)
            all_rows_for_consolidation.sort(key=lambda r: parse_bit_ts(r[3], r[4]) or datetime.min)
        
        for row in all_rows_for_consolidation:
            consolidated_ws.append(row)
        apply_status_formatting(consolidated_ws)
        autoformat(consolidated_ws)

    autoformat(ws_summary); out_path = out_dir / f"BitStatus_Results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(out_path)
    return out_path, rows_added

class BitWorker(QThread):
    log_msg = Signal(str); progress = Signal(int, int); result_ready = Signal(dict); failed = Signal(str)
    def __init__(self, mode: str, selection: List[str], out_dir: str, filter_date: Optional[str] = None, filter_start: Optional[str] = None, filter_end: Optional[str] = None, consolidate: bool = False, sort_chronologically: bool = True):
        super().__init__()
        self.mode, self.selection, self.out_dir = mode, selection, Path(out_dir)
        self.filter_date, self.filter_start, self.filter_end, self.consolidate = filter_date, filter_start, filter_end, consolidate
        self.sort_chronologically = sort_chronologically
        self._cancel = False
    def request_cancel(self): self._cancel = True
    def _collect_files(self) -> List[Path]:
        _td = _safe_temp_root()
        _z = lambda p: _extract_logs_from_zip(p, _td, 'user') if p.suffix.lower() == '.zip' else []
        if self.mode == 'folder':
            f = Path(self.selection[0]); l = []
            for p in f.rglob('*'):
                if p.is_file() and 'user' in p.name.lower() and p.suffix.lower() == '.log': l.append(p)
                elif p.is_file() and p.suffix.lower() == '.zip': l.extend(_z(p))
            return sorted(l)
        if self.mode == 'files':
            l = []
            for s in self.selection:
                p = Path(s)
                if p.is_file() and p.suffix.lower() == '.zip': l.extend(_z(p))
                elif p.is_file() and 'user' in p.name.lower() and p.suffix.lower() == '.log': l.append(p)
            return sorted(l)
        if self.mode == 'single':
            p = Path(self.selection[0])
            if p.suffix.lower() == '.zip': return _extract_logs_from_zip(p, _td, 'user')
            return [p] if p.is_file() else []
        return []
    def run(self):
        try:
            files = self._collect_files()
            if not files: self.failed.emit("No files to process."); return
            self.out_dir.mkdir(parents=True, exist_ok=True)
            out_path, cnt = process_bit_files(files, self.out_dir, self.filter_date, self.filter_start, self.filter_end, self.consolidate, self.progress.emit, lambda: self._cancel, sort_chronologically=self.sort_chronologically, log_callback=self.log_msg.emit)
            if out_path: 
                self.result_ready.emit({'report_path': str(out_path), 'count': cnt})
                if cnt == 0:
                    self.log_msg.emit("<b style='color:orange'>Warning: 0 rows found. Check your AM/PM and Date settings.</b>")
        except Exception as ex: self.failed.emit(str(ex))

# -----------------------------------------------------------------------------
# ERROR LOG Logic
# -----------------------------------------------------------------------------

ERR_HDR_PROGRAM_RE = re.compile(r'Program\s+"(?P<program>[^"]+)"\s+at address\s+(?P<address>\d+)\s+on\s+(?P<com>\S+)', re.IGNORECASE)
ERR_PC_DATE_RE = re.compile(r'PC Date\s*:(?P<pcdate>.*)', re.IGNORECASE)
YEAR_CHANGE_RE = re.compile(r"Year is changing from '?(?P<from_y>\d+) to '?(?P<to_y>\d+)", re.IGNORECASE)
ERR_ENTRY_RE = re.compile(r'^\[(?P<level>\w+)\]\s+(?P<md>\d{2}/\d{2})\s+(?P<time>\d{2}:\d{2}:\d{2}(?:\.\d+)?)\s+(?P<ampm>AM|PM)\s+(?P<message>.*?)(?:\s+\[(?P<codes>.*?)\])?$', re.IGNORECASE)

def parse_pcdate_year(pcd: str) -> Optional[int]:
    m = re.search(r'\b(20\d{2})\b', pcd)
    return int(m.group(1)) if m else None

def extract_station(prog: Optional[str]) -> str:
    if not prog: return "-"
    return prog.split('.')[0] if '.' in prog else prog

def err_classify_failure(msg: str) -> str:
    m = msg.lower()
    # --- System-level resets ---
    if 'on-line sync reset' in m:           return 'On-Line Sync Reset'
    if 'off-line cps sync reset' in m:      return 'Off-Line CPS Sync Reset'
    if 'sync quarantine transition' in m:   return 'Sync Quarantine Transition'
    if 'watchdog hit' in m or 'system reset' in m: return 'System Reset - Watchdog'
    if 'year is changing' in m:             return 'Year Change Event'
    # --- Memory / data errors ---
    if 'double store error' in m:           return 'Double Store Error'
    if 'data verification' in m or 'data mismatch' in m: return 'Data Verification Error'
    # --- Executive / loading ---
    if 'executive loading' in m or 'loading error' in m: return 'Executive Loading Error'
    # --- Kill bits ---
    if 'kill bit' in m:                     return 'Kill Bit (Application)'
    # --- Internal vector errors ---
    if 'internal vector' in m:
        if 'bus_error' in m or 'bus error' in m: return 'Internal Vector - Bus Error'
        if 'spurious' in m:                      return 'Internal Vector - Spurious'
        return 'Internal Vector Error'
    # --- Logic queue ---
    if 'logic queue overflow' in m:         return 'Logic Queue Overflow'
    # --- Sync board / link ---
    if 'sync i/o board' in m or 'sync board' in m:
        if 'type error' in m:               return 'Sync I/O Board Type Error'
        if 'echo error' in m:               return 'Sync I/O Board Echo Error'
        return 'Sync Board Error'
    if 'sync link' in m and 'discrepan' in m: return 'Sync Link Discrepancy'
    # --- Generic I/O board errors (catches NVIN32, OUT16, etc.) ---
    if 'output failure' in m:               return 'Output Failure'
    if 'type error' in m:                   return 'Type Error'
    if 'echo error' in m:                   return 'Echo Error'
    # --- Smart structural fallback ---
    # Most Microlok II messages follow: "Board/Component description - Error Qualifier [hex codes]"
    # Extract the qualifier part (after last ' - ') and strip board-specific noise.
    if ' - ' in msg:
        qualifier = msg.split(' - ')[-1].strip()
        # Strip trailing hex codes like "0E01", "VEC...", "032..."
        qualifier = re.sub(r'\s+[0-9A-Fa-f]{3,}\S*\s*$', '', qualifier).strip()
        # Strip leading hex/numeric prefixes (e.g. "0x1A: some error")
        qualifier = re.sub(r'^[0-9A-Fa-fx:]+\s+', '', qualifier).strip()
        if qualifier and len(qualifier) > 3:
            return qualifier[:55]  # cap length for legend readability
    # Last resort: first 40 chars
    return (msg[:40] + '...') if len(msg) > 40 else msg


def _cardfile_sort_key(cf: str) -> Tuple[str, int]:
    m = re.search(r'(\d+)', cf)
    return (re.sub(r'\d+', '', cf), int(m.group(1)) if m else 0)

def _normalize_y_dt(dt: datetime, scale: str) -> datetime:
    if scale == 'Date': return dt.replace(hour=0, minute=0, second=0, microsecond=0)
    if scale == 'Hour': return dt.replace(minute=0, second=0, microsecond=0)
    if scale == 'Minute': return dt.replace(second=0, microsecond=0)
    return dt

def pick_auto_scale(t_min: datetime, t_max: datetime) -> str:
    diff = t_max - t_min
    if diff > timedelta(days=30): return 'Month'
    if diff > timedelta(days=2): return 'Date'
    if diff > timedelta(hours=2): return 'Hour'
    return 'Minute'

def err_save_xlsx(rows: List[Dict[str, Any]], out_path: Path):
    wb = Workbook(); ws = wb.active; ws.title = "Errors"
    if not rows: return
    
    # User requested removing "Station", "Address", "Com"
    # Original: ["File Path", "File Name", "Cardfile", "Program", "Station", "Address", "Com", "Level", "MonthDay", "Time", "Failure Type", "Message", "Codes"]
    headers = ["File Path", "File Name", "Cardfile", "Program", "Level", "MonthDay", "Time", "Failure Type", "Message", "Codes"]
    ws.append(headers)
    
    # Style Header
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    for r in rows:
        ws.append([
            r.get('file_path',''), 
            r.get('file_name',''), 
            r.get('cardfile',''), 
            r.get('program',''), 
            # r.get('station',''), # Removed 
            # r.get('address',''), # Removed
            # r.get('com',''),     # Removed
            r.get('level',''), 
            r.get('month_day',''), 
            r.get('time',''), 
            r.get('failure_type',''), 
            r.get('message',''), 
            r.get('codes','')
        ])
    autoformat(ws); wb.save(out_path)

def extract_error_cardfile(path: Path) -> str:
    # Smart cardfile extraction: If filename is generic "Errors...", look deeper
    stem = path.stem
    if stem.lower().startswith("errors"):
        # Try to find something like A1, B2 in the filename
        m_cf = re.search(r'\b([AB]\d+)\b', stem, re.I)
        return m_cf.group(1).upper() if m_cf else stem
    else:
        return stem.split()[0] if stem else ""

def err_iter_records(path: Path) -> Iterable[dict]:
    cardfile = extract_error_cardfile(path)

    program = address = com = None
    pc_date = None; cy = None
    with path.open('r', encoding='utf-8', errors='ignore') as fh:
        for line in (l.rstrip() for l in fh):
            mh = ERR_HDR_PROGRAM_RE.search(line); (program := mh.group('program'), address := mh.group('address'), com := mh.group('com')) if mh else None
            mp = ERR_PC_DATE_RE.search(line); (pc_date := mp.group('pcdate').strip(), cy := parse_pcdate_year(pc_date) or cy) if mp else None
            
            # Detect year change in error logs
            myc = YEAR_CHANGE_RE.search(line)
            if myc:
                # In reverse chronological error logs, we transition to the EARLIER year
                y_val = int(myc.group('from_y'))
                cy = y_val + 2000 if y_val < 100 else y_val

            m = ERR_ENTRY_RE.match(line)
            if not m: continue
            h, mi, s = m.group('time').strip().split(':'); hr = int(h); hr += 12 if m.group('ampm').upper() == 'PM' and hr != 12 else 0; hr = 0 if m.group('ampm').upper() == 'AM' and hr == 12 else hr
            mo, d = map(int, m.group('md').split('/')); yr = cy or datetime.now().year
            try: dt = datetime(yr, mo, d, hr, int(mi), int(float(s)), int((float(s) % 1) * 1000000))
            except Exception: continue
            
            # Extract codes (already handled by single group)
            codes_val = (m.group('codes') or '').strip()
            
            yield {
                'file_path': str(path.parent.name),
                'file_name': str(path.name),
                'source_file': str(path.parent.name), 
                'cardfile': str(cardfile), 
                'program': str(program or ""), 
                'station': str(extract_station(program)), 
                'address': str(address or ""), 
                'com': str(com or ""), 
                'pc_date': str(pc_date or ""), 
                'level': str(m.group('level').lower()), 
                'month_day': str(m.group('md')), 
                'time': str(f"{m.group('time').strip()} {m.group('ampm')}").replace('\t', ' '), 
                'time_dt': dt, 
                'message': str(m.group('message').strip()).replace('\t', ' '), 
                'codes': str(codes_val or "").replace('\t', ' '), 
                'failure_type': str(err_classify_failure(m.group('message').strip())).replace('\t', ' ')
            }


def apply_error_filters(rows: List[dict], station: Optional[str] = None, time_filter: Optional[dict] = None) -> List[dict]:
    """
    Filter error rows by station and time range.
    """
    # 1. Filter by Station
    if station and station != 'All':
        rows = [r for r in rows if r.get('station') == station]

    # 2. Filter by Time
    if not time_filter:
        return rows

    # Helpers for time matching
    def _coerce_to_pytime(obj) -> Optional[dtime]:
        if obj in (None, 'All'): return None
        try:
            if hasattr(obj, 'toPython'): return obj.toPython()
        except Exception: pass
        try: return dtime(obj.hour, obj.minute, obj.second)
        except Exception: pass
        return obj if isinstance(obj, dtime) else None

    def _coerce_to_pydate(obj):
        if obj in (None, 'All'):
            return None
        try:
            if hasattr(obj, 'toPython'):
                val = obj.toPython()
                return val.date() if isinstance(val, datetime) else val
        except Exception:
            pass
        if isinstance(obj, datetime):
            return obj.date()
        if hasattr(obj, 'year') and hasattr(obj, 'month') and hasattr(obj, 'day'):
            try:
                return datetime(int(obj.year), int(obj.month), int(obj.day)).date()
            except Exception:
                pass
        if isinstance(obj, str):
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y", "%d/%m/%Y", "%m/%d/%Y"):
                try:
                    return datetime.strptime(obj, fmt).date()
                except Exception:
                    pass
        return None

    yr = time_filter.get('year', 'All')
    mo = time_filter.get('month', 'All')
    dy = time_filter.get('day', 'All')
    t_from = time_filter.get('t_from', 'All')
    t_to = time_filter.get('t_to', 'All')
    d_from = _coerce_to_pydate(time_filter.get('date_from'))
    d_to = _coerce_to_pydate(time_filter.get('date_to'))
    if d_from and d_to and d_from > d_to:
        d_from, d_to = d_to, d_from
    
    # Parse times once
    low = _coerce_to_pytime(t_from) if t_from != 'All' else None
    high = _coerce_to_pytime(t_to) if t_to != 'All' else None

    # Parse date parts to int if possible
    try: yr = int(yr) if yr != 'All' else 'All'
    except: yr = 'All'
    try: mo = int(mo) if mo != 'All' else 'All'
    except: mo = 'All'
    try: dy = int(dy) if dy != 'All' else 'All'
    except: dy = 'All'

    filtered = []
    for r in rows:
        dt = r.get('time_dt')
        if not dt: continue

        if d_from and dt.date() < d_from:
            continue
        if d_to and dt.date() > d_to:
            continue
        
        # Date Level Checks
        if yr != 'All' and dt.year != yr: continue
        if mo != 'All' and dt.month != mo: continue
        if dy != 'All' and dt.day != dy: continue
        
        # Time Range Checks
        if low and high and low != high:
            t = dt.time()
            if low <= high:
                if not (low <= t <= high): continue
            else:
                # Overnight wrap (e.g. 11PM to 2AM)
                if not (t >= low or t <= high): continue
        
        filtered.append(r)
    
    return filtered

class ErrorWorker(QThread):
    log_msg = Signal(str); progress = Signal(int, int); result_ready = Signal(dict); failed = Signal(str)
    def __init__(self, mode: str, selection: List[str], out_dir: str, include_events: bool, y_scale: str, 
                 station: str = "All", time_filter: Optional[dict] = None, sort_chrono: bool = False):
        super().__init__()
        self.mode, self.selection, self.out_dir = mode, selection, Path(out_dir)
        self.include_events, self.y_scale = include_events, y_scale
        self.station, self.time_filter = station, time_filter
        self.sort_chrono = sort_chrono
        self._cancel = False
    def request_cancel(self): self._cancel = True
    def _collect_files(self) -> List[Path]:
        l = []; self._td = _safe_temp_root()
        _z = lambda p: _extract_logs_from_zip(p, self._td, 'error') if p.suffix.lower() == '.zip' else []
        if self.mode == 'folder':
            for p in sorted(Path(self.selection[0]).rglob('*')):
                if p.is_file() and 'error' in p.name.lower() and p.suffix.lower() == '.log': l.append(p)
                elif p.is_file() and p.suffix.lower() == '.zip': l.extend(_z(p))
        elif self.mode == 'files':
            for f in self.selection:
                p = Path(f)
                if 'error' in p.name.lower() and p.suffix.lower() == '.log': l.append(p)
                elif p.suffix.lower() == '.zip': l.extend(_z(p))
        return sorted(l)
    def run(self):
        try:
            files = self._collect_files()
            if not files: self.failed.emit("No files found."); return
            
            all_cardfiles = sorted(list({extract_error_cardfile(f) for f in files}), key=_cardfile_sort_key)
            
            self.out_dir.mkdir(parents=True, exist_ok=True); rows = []
            for i, f in enumerate(files, start=1):
                if self._cancel: break
                self.progress.emit(i, len(files))
                self.log_msg.emit(f"Analyzing: {f.name}")
                rows.extend(err_iter_records(f))
            if not rows: self.failed.emit("No entries."); return
            self.log_msg.emit(f"Extracted {len(rows)} raw entries.")
            
            # Apply Filters
            filtered_rows = apply_error_filters(rows, self.station, self.time_filter)
            self.log_msg.emit(f"After filtering: {len(filtered_rows)} entries.")

            if not filtered_rows: self.failed.emit("No entries matched your filters."); return
            
            # Apply Chronological Sort if requested
            if self.sort_chrono:
                filtered_rows.sort(key=lambda x: x.get('time_dt') or datetime.min)
                self.log_msg.emit("Sorted output chronologically.")
            
            out = self.out_dir / f"parsed_errors_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            err_save_xlsx(filtered_rows, out)
            
            # Generate initial chart
            # We pass filtered_rows, but we also pass all_cardfiles so the X-axis can be stable or show context if needed
            # Also pass station for the title.
            png_bytes = err_plot_time_chart(
                filtered_rows, 
                y_scale=self.y_scale, 
                station=self.station,
                include_events=self.include_events, 
                all_cardfiles=all_cardfiles,
                time_filter=self.time_filter
            )
            
            res = {
                'report_path': str(out),
                'error_rows': filtered_rows,
                'card_files': all_cardfiles,
                'chart_png': png_bytes
            }
            self.result_ready.emit(res)
        except Exception as ex: self.failed.emit(str(ex))

# -----------------------------------------------------------------------------
# Global Upload Logic
# -----------------------------------------------------------------------------

class GlobalWorker(QThread):
    log_msg = Signal(str); progress = Signal(int, int); result_ready = Signal(dict); failed = Signal(str)
    def __init__(self, mode: str, selection: List[str], out_dir: str, rules_path: Optional[str], export_proposed: bool, sort_chronologically: bool = True):
        super().__init__()
        self.mode, self.selection, self.out_dir, self.rules_path, self.export_proposed = mode, selection, Path(out_dir), rules_path, export_proposed
        self.sort_chronologically = sort_chronologically
        self._cancel = False
    def request_cancel(self): self._cancel = True
    def _collect_files(self) -> List[Path]:
        l = []; self._td = _safe_temp_root()
        if self.mode == 'folder':
            for p in sorted(Path(self.selection[0]).rglob('*.log')): l.append(p)
        elif self.mode == 'files':
            for f in self.selection:
                p = Path(f); (l.append(p) if p.suffix.lower() == '.log' else None)
        return sorted(l)
    def run(self):
        try:
            files = self._collect_files()
            if not files: self.failed.emit("No files to process."); return
            self.out_dir.mkdir(parents=True, exist_ok=True)
            
            # Categorize files
            event_files = [f for f in files if 'event' in f.name.lower()]
            bit_files = [f for f in files if 'user' in f.name.lower()]
            error_files = [f for f in files if 'error' in f.name.lower()]
            
            total = len(event_files) + len(bit_files) + len(error_files)
            if total == 0: self.failed.emit("No recognizable log files found."); return
            
            self.log_msg.emit(f"Global: Groups: {len(event_files)} Event, {len(bit_files)} User, {len(error_files)} Error.")
            processed_count = 0
            
            rules = load_rules(self.rules_path)
            
            # Helper for nested progress
            def make_prog(current, group_total, offset, overall_total):
                p = int(100 * (offset + current) / max(1, overall_total))
                self.progress.emit(p, 100)

            results = []
            if event_files:
                self.log_msg.emit("Processing Event logs...")
                res, cnt = process_event_files(event_files, self.out_dir, rules, lambda c, t: make_prog(c, t, processed_count, total), lambda: self._cancel)
                if res: results.append(str(res)); self.log_msg.emit(f"Event summary: {res.name} ({cnt} entries)")
                processed_count += len(event_files)
            
            if self._cancel: return
            
            if bit_files:
                self.log_msg.emit("Processing User logs...")
                res, cnt = process_bit_files(bit_files, self.out_dir, None, None, None, False, lambda c, t: make_prog(c, t, processed_count, total), lambda: self._cancel, sort_chronologically=self.sort_chronologically)
                if res: results.append(str(res)); self.log_msg.emit(f"User summary: {res.name} ({cnt} rows)")
                processed_count += len(bit_files)

            if self._cancel: return
            
            if error_files:
                self.log_msg.emit("Processing Error logs...")
                rows = []
                for i, f in enumerate(error_files, start=1):
                    if self._cancel: break
                    make_prog(i, len(error_files), processed_count, total)
                    rows.extend(err_iter_records(f))
                if rows:
                    out = self.out_dir / f"parsed_errors_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    err_save_xlsx(rows, out)
                    results.append(str(out))
                    self.log_msg.emit(f"Error summary: {out.name} ({len(rows)} entries)")
                processed_count += len(error_files)

            if not results: self.failed.emit("No results generated.")
            else: self.result_ready.emit({'report_path': None, 'reports': results, 'msg': f"Global done. {len(results)} reports generated."})
        except Exception as ex: self.failed.emit(str(ex))

# ----------------------------- Y-axis helpers -------------------------------
def _normalize_y_dt(dt: datetime, scale: str) -> datetime:
    if scale == 'Year': return datetime(dt.year, 1, 1)
    if scale == 'Month': return datetime(dt.year, dt.month, 1)
    if scale == 'Date': return datetime(dt.year, dt.month, dt.day)
    if scale == 'Hour': return datetime(dt.year, dt.month, dt.day, dt.hour)
    if scale == 'Minute': return datetime(dt.year, dt.month, dt.day, dt.hour, dt.minute)
    if scale == 'Second': return datetime(dt.year, dt.month, dt.day, dt.hour, dt.minute, dt.second)
    return dt

def _adaptive_locator_for_span(min_dt: datetime, max_dt: datetime, scale: str, target_ticks: int = 200):
    span = max(max_dt - min_dt, timedelta(seconds=1))
    if scale == 'Second':
        total = max(1, math.ceil(span.total_seconds()))
        interval = max(1, math.ceil(total / target_ticks))
        return mdates.SecondLocator(interval=interval), mdates.DateFormatter('%H:%M:%S')
    if scale == 'Minute':
        total = math.ceil(span.total_seconds() / 60)
        interval = max(1, math.ceil(total / target_ticks))
        return mdates.MinuteLocator(interval=interval), mdates.DateFormatter('%H:%M')
    if scale == 'Hour':
        total = math.ceil(span.total_seconds() / 3600)
        interval = max(1, math.ceil(total / target_ticks))
        return mdates.HourLocator(interval=interval), mdates.DateFormatter('%H:00')
    if scale == 'Date':
        total = max(1, span.days)
        interval = max(1, math.ceil(total / target_ticks))
        return mdates.DayLocator(interval=interval), mdates.DateFormatter('%d-%b-%Y')
    if scale == 'Month':
        months = max(1, math.ceil(span.days / 30.0))
        interval = max(1, math.ceil(months / target_ticks))
        return mdates.MonthLocator(interval=interval), mdates.DateFormatter('%b %Y')
    if scale == 'Year':
        years = max(1, math.ceil(span.days / 365.0))
        base = max(1, math.ceil(years / target_ticks))
        return mdates.YearLocator(base=base), mdates.DateFormatter('%Y')
    loc = mdates.AutoDateLocator(minticks=5, maxticks=15)
    return loc, mdates.ConciseDateFormatter(loc)

def pick_auto_scale(ymin: datetime, ymax: datetime) -> str:
    span = max(ymax - ymin, timedelta(seconds=1))
    if span <= timedelta(hours=2): return 'Second'
    if span <= timedelta(hours=12): return 'Minute'
    if span <= timedelta(days=3): return 'Hour'
    if span <= timedelta(days=120): return 'Date'
    if span <= timedelta(days=730): return 'Month'
    return 'Year'

def _cardfile_sort_key(cf: str):
    s = str(cf).replace('_', '').strip()
    m = re.match(r'^([A-Za-z]+)\s*0*([0-9]+)$', s)
    if m:
        letters, num = m.group(1).upper(), int(m.group(2))
        return (num, letters)
    return (float('inf'), str(cf))

# ------------------------------- Plotter ------------------------------------
def err_plot_time_chart(rows: List[dict], *,
    y_scale: str = 'Auto', 
    station: Optional[str] = None,
    include_events: bool = False,
    dynamic_width: bool = True,
    time_filter: Optional[dict] = None,
    all_cardfiles: Optional[List[str]] = None, # New param
    title_extras: Optional[str] = None,
    max_ticks: int = 1000,
    return_png_bytes: bool = True,
    save_path: Optional[Path] = None
) -> Optional[bytes]:
    """
    Renders the time chart and either returns PNG bytes for preview (default) or saves to save_path.
    """
    # Filter by level
    data = rows if include_events else [r for r in rows if r.get('level') == 'error']
    
    # Apply Time/Station Filters using the shared helper
    # logic is robust to types now
    if station or time_filter:
        data = apply_error_filters(data, station, time_filter)

    if not data: return None

    # X categories (cardfiles)
    if all_cardfiles:
        cardfiles = all_cardfiles
    else:
        cardfiles = sorted({r['cardfile'] for r in data if r.get('cardfile')}, key=_cardfile_sort_key)
    
    x_map = {cf: i for i, cf in enumerate(cardfiles)} # ... (rest of function)
    
    # Failure types (legend)
    ftypes = []
    for r in data:
        ft = r.get('failure_type', 'Other')
        if ft not in ftypes: ftypes.append(ft)
    
    marker_cycle = ['o', 's', '^', 'D', 'v', 'P', '*', 'X', 'h', '8', '<', '>']
    color_cycle = plt.get_cmap('tab10').colors
    markers = {ft: marker_cycle[i % len(marker_cycle)] for i, ft in enumerate(ftypes)}
    colors = {ft: color_cycle[i % len(color_cycle)] for i, ft in enumerate(ftypes)}
    
    # Dynamic figure width
    n_x = max(1, len(cardfiles))
    fig_w = max(8.0, min(36.0, 8.0 + 0.6 * (n_x - 5))) if dynamic_width else 10.0
    
    # Pre-calculate rotation for margins
    max_len = max([len(str(c)) for c in cardfiles]) if cardfiles else 0
    do_rotate = len(cardfiles) > 5 or max_len > 8
    rotation = 45 if len(cardfiles) <= 15 else 90
    align = 'right' if rotation == 45 else 'center'
    
    # Figure & Axes
    fig, ax = plt.subplots(figsize=(fig_w, 8)) # Increased height from 6 to 8
    fig.patch.set_facecolor('white'); ax.set_facecolor('white')
    # Increased margins for labels - Dynamic based on rotation
    bottom_margin = 0.35 if do_rotate else 0.15
    if rotation == 90: bottom_margin = 0.45
    
    # Reserve space on the right for the legend
    fig.subplots_adjust(left=0.10, right=0.75, top=0.90, bottom=bottom_margin)
    
    y_dts_raw = [r['time_dt'] for r in data]
    ymin_raw, ymax_raw = min(y_dts_raw), max(y_dts_raw)
    if y_scale == 'Auto': y_scale = pick_auto_scale(ymin_raw, ymax_raw)
    
    def _norm(dt: datetime): return _normalize_y_dt(dt, y_scale)
    
    y_dts = [_norm(dt) for dt in y_dts_raw]
    unique_y_dts = sorted(list(set(y_dts)))

    def coerce_to_datetime(obj) -> Optional[datetime]:
        if not obj or obj == 'All': return None
        try:
            if hasattr(obj, 'toPython'):
                val = obj.toPython()
                if isinstance(val, datetime): return val
                if hasattr(val, 'year') and hasattr(val, 'month') and hasattr(val, 'day'):
                    return datetime(val.year, val.month, val.day)
        except: pass
        if isinstance(obj, datetime): return obj
        if hasattr(obj, 'year') and hasattr(obj, 'month') and hasattr(obj, 'day'):
            try: return datetime(int(obj.year), int(obj.month), int(obj.day))
            except: pass
        if isinstance(obj, str):
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y", "%d/%m/%Y", "%m/%d/%Y"):
                try: return datetime.strptime(obj, fmt)
                except: pass
        return None

    custom_start = coerce_to_datetime(time_filter.get('date_from')) if time_filter else None
    custom_end = coerce_to_datetime(time_filter.get('date_to')) if time_filter else None

    # Include custom date bounds if specified
    if custom_start:
        norm_start = _norm(custom_start)
        if norm_start not in unique_y_dts:
            unique_y_dts.append(norm_start)
    if custom_end:
        norm_end = _norm(custom_end)
        if norm_end not in unique_y_dts:
            unique_y_dts.append(norm_end)
    unique_y_dts = sorted(list(set(unique_y_dts)))

    use_discrete_y = len(unique_y_dts) <= 60

    if use_discrete_y:
        y_map = {dt: i for i, dt in enumerate(unique_y_dts)}
        for ft in ftypes:
            sub = [r for r in data if r.get('failure_type') == ft]
            xs = [x_map.get(r['cardfile'], -1) for r in sub]
            ys = [y_map[_norm(r['time_dt'])] for r in sub]
            ax.scatter(xs, ys, marker=markers[ft], color=colors[ft], edgecolor='black', linewidths=0.5, s=60, label=ft)
        
        # X ticks
        ax.set_xticks(list(x_map.values()))
        ax.set_xticklabels(cardfiles, rotation=rotation, ha=align, fontsize=8 if len(cardfiles) > 10 else 9)
        ax.set_xlabel('Cardfile', color='black')
        
        # Y ticks and limits (Discrete Mode)
        ax.set_yticks(list(range(len(unique_y_dts))))
        labels = []
        for dt in unique_y_dts:
            if y_scale == 'Second': lbl = dt.strftime('%H:%M:%S')
            elif y_scale == 'Minute': lbl = dt.strftime('%H:%M')
            elif y_scale == 'Hour': lbl = dt.strftime('%H:00')
            elif y_scale == 'Date': lbl = dt.strftime('%d-%b-%Y')
            elif y_scale == 'Month': lbl = dt.strftime('%b %Y')
            elif y_scale == 'Year': lbl = dt.strftime('%Y')
            else: lbl = dt.strftime('%Y-%m-%d %H:%M:%S')
            labels.append(lbl)
        ax.set_yticklabels(labels)
        ax.tick_params(axis='y', labelsize=8)
        ax.set_ylim(-0.5, len(unique_y_dts) - 0.5)
    else:
        # Continuous (Chronological) plot logic
        for ft in ftypes:
            sub = [r for r in data if r.get('failure_type') == ft]
            xs = [x_map.get(r['cardfile'], -1) for r in sub]
            ys = [mdates.date2num(_norm(r['time_dt'])) for r in sub]
            ax.scatter(xs, ys, marker=markers[ft], color=colors[ft], edgecolor='black', linewidths=0.5, s=60, label=ft)
            
        # X ticks
        ax.set_xticks(list(x_map.values()))
        ax.set_xticklabels(cardfiles, rotation=rotation, ha=align, fontsize=8 if len(cardfiles) > 10 else 9)
        ax.set_xlabel('Cardfile', color='black')
        
        # Y axis
        ymin, ymax = min(y_dts), max(y_dts)
        if custom_start:
            ymin = min(ymin, _norm(custom_start))
        if custom_end:
            ymax = max(ymax, _norm(custom_end))

        if ymin == ymax:
            # Expand slightly to avoid singular transformation warning
            if y_scale == 'Second': delta = timedelta(seconds=5)
            elif y_scale == 'Minute': delta = timedelta(minutes=1)
            elif y_scale == 'Hour': delta = timedelta(hours=1)
            elif y_scale == 'Date': delta = timedelta(days=1)
            elif y_scale == 'Month': delta = timedelta(days=30)
            elif y_scale == 'Year': delta = timedelta(days=365)
            else: delta = timedelta(minutes=5)
            ymin -= delta; ymax += delta
        else:
            # Add padding/margin to prevent boundary markers from being clipped
            span = ymax - ymin
            pad_seconds = span.total_seconds() * 0.05
            
            # Ensure a minimum padding based on the Y-scale
            if y_scale == 'Second': min_pad = 2
            elif y_scale == 'Minute': min_pad = 10
            elif y_scale == 'Hour': min_pad = 600
            elif y_scale == 'Date': min_pad = 86400
            elif y_scale == 'Month': min_pad = 86400 * 10
            elif y_scale == 'Year': min_pad = 86400 * 30
            else: min_pad = 300
            
            padding = timedelta(seconds=max(pad_seconds, min_pad))
            ymin -= padding; ymax += padding

        ax.set_ylim(mdates.date2num(ymin), mdates.date2num(ymax))
        loc, fmt = _adaptive_locator_for_span(ymin, ymax, y_scale, target_ticks=40)
        
        # Tick Count safety to prevent freeze
        try:
            t_v = list(loc.tick_values(ymin, ymax))
            if len(t_v) > 500:
                loc = mdates.AutoDateLocator(minticks=5, maxticks=12)
                fmt = mdates.AutoDateFormatter(loc)
                t_v = list(loc.tick_values(ymin, ymax))
        except:
            t_v = []

        ymin_norm, ymax_norm = min(y_dts), max(y_dts)
        if custom_start:
            ymin_norm = min(ymin_norm, _norm(custom_start))
        if custom_end:
            ymax_norm = max(ymax_norm, _norm(custom_end))

        y_min_num = mdates.date2num(ymin_norm)
        y_max_num = mdates.date2num(ymax_norm)

        if len(t_v) > 0:
            if len(t_v) > 1:
                spacing = t_v[1] - t_v[0]
                threshold = 0.7 * spacing
            else:
                if y_scale == 'Second': threshold = 1.5 / 86400.0
                elif y_scale == 'Minute': threshold = 45.0 / 86400.0
                elif y_scale == 'Hour': threshold = 2700.0 / 86400.0
                elif y_scale == 'Date': threshold = 0.8
                elif y_scale == 'Month': threshold = 12.0
                elif y_scale == 'Year': threshold = 75.0
                else: threshold = 0.8
            
            filtered_ticks = []
            for t in t_v:
                if abs(t - y_min_num) >= threshold and abs(t - y_max_num) >= threshold:
                    filtered_ticks.append(t)
            
            filtered_ticks.append(y_min_num)
            if y_min_num != y_max_num:
                filtered_ticks.append(y_max_num)
            
            filtered_ticks.sort()
            ax.set_yticks(filtered_ticks)
        else:
            ax.set_yticks([y_min_num, y_max_num] if y_min_num != y_max_num else [y_min_num])

        ax.yaxis.set_major_formatter(fmt)
        ax.tick_params(axis='y', labelsize=8)
    
    ylabel = {'Year':'Year','Month':'Month','Date':'Date','Hour':'Hour','Minute':'Minute','Second':'Second'}.get(y_scale, 'Time')
    ax.set_ylabel(ylabel, labelpad=12, color='black')
    ax.grid(True, which='both', linestyle='--', alpha=0.3)
    
    title_station = f"Station: {station}" if station and station != 'All' else "Station: All"
    sub = title_extras or ""
    ax.set_title(f'Error Events by Cardfile vs {y_scale}\n{title_station}{" • " + sub if sub else ""}', color='black')
    
    # Place legend outside to the right
    leg = fig.legend(title='Failure Type', loc='center left', bbox_to_anchor=(0.76, 0.5), frameon=True, fontsize=9)
    if leg:
        plt.setp(leg.get_title(), fontsize=10, color='black')
        for text in leg.get_texts(): text.set_color('black')
    
    if save_path:
        fig.savefig(save_path, dpi=150, bbox_inches='tight')
        plt.close(fig); return None
    
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    plt.close(fig); return buf.getvalue()
